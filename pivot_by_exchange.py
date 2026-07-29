"""
Pivot the long TradFi-perp table into an underlying x exchange volume matrix.

Input  : long CSV, one row per contract (output of tradfi_perps_collect.py,
         or your existing sheet exported to CSV)
Output : XLSX with two tabs
           Volume Matrix - rows = underlying, cols = exchange, cells = 24h vol USD
           Contract Detail - the original long table, kept for traceability

    python pivot_by_exchange.py universe.csv --out tradfi_matrix.xlsx

Hyperliquid note: xyz:/flx:/cash:/para:/vntl: are separate builder-deployed
markets with independent order books, so they are kept as separate columns by
default. Pass --merge-hl to collapse them into one Hyperliquid column.
"""

import argparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
GROUP_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="BFBFBF")


def split_hyperliquid(df):
    """Hyperliquid pairs look like 'xyz:NVDA'. Promote the namespace into the
    exchange name so each builder market gets its own column."""
    is_hl = df["Exchange"].str.contains("hyperliquid", case=False, na=False)
    ns = df.loc[is_hl, "Pair"].str.extract(r"^([a-z]+):", expand=False)
    df.loc[is_hl & ns.notna(), "Exchange"] = "Hyperliquid " + ns[ns.notna()]
    return df


def clean_volume(s):
    """'1,899,429,449.00' / 'N/A' / '' -> float or NaN."""
    return pd.to_numeric(
        s.astype(str).str.replace(r"[,$\s]", "", regex=True)
         .replace({"N/A": None, "": None, "nan": None, "--": None}),
        errors="coerce",
    )


def build(df, merge_hl=False):
    if not merge_hl:
        df = split_hyperliquid(df)

    df["_vol"] = clean_volume(df["24h Volume (USDT)"])

    # Group label: whichever identifier the input actually carries.
    # Exchange-API output has Ticker; CMC market-pairs output has only Underlying.
    for col in ("Underlying", "Ticker", "Pair"):
        if col in df.columns:
            df["_group"] = df[col] if "_group" not in df else df["_group"].fillna(df[col])
    df = df[df["_group"].notna()]

    # sum, not first: one exchange can list several contracts on one underlying
    # (Binance STX + STXX, tokenized xStock variants, etc.)
    matrix = df.pivot_table(index=["_group", "Classification"]
                            if "Classification" in df.columns else ["_group"],
                            columns="Exchange", values="_vol",
                            aggfunc="sum", dropna=False)

    # a 2-level index with dropna=False emits the full Underlying x Classification
    # cross-product; keep only combinations that actually have a contract
    matrix = matrix.dropna(how="all")

    matrix = matrix.reindex(sorted(matrix.columns), axis=1).reset_index()
    matrix = matrix.rename(columns={"_group": "Underlying"})

    # order rows by asset class, then by total volume desc
    exch_cols = [c for c in matrix.columns
                 if c not in ("Underlying", "Classification")]
    matrix["_sort"] = matrix[exch_cols].sum(axis=1)
    sort_keys = (["Classification", "_sort"] if "Classification" in matrix.columns
                 else ["_sort"])
    matrix = (matrix.sort_values(sort_keys, ascending=[True, False]
                                 if len(sort_keys) == 2 else [False])
              .drop(columns="_sort").reset_index(drop=True))
    return matrix, exch_cols, df


def write(matrix, exch_cols, detail, out):
    with pd.ExcelWriter(out, engine="openpyxl") as xl:
        matrix.to_excel(xl, sheet_name="Volume Matrix", index=False,
                        startrow=1)
        detail.drop(columns=[c for c in ("_vol", "_group") if c in detail],
                    errors="ignore").to_excel(
            xl, sheet_name="Contract Detail", index=False)

    wb = load_workbook(out)
    ws = wb["Volume Matrix"]
    n_rows = len(matrix)
    label_cols = len(matrix.columns) - len(exch_cols)

    ws["A1"] = "24h trading volume (USDT) by underlying and exchange"
    ws["A1"].font = Font(name=FONT, size=12, bold=True)

    # Venues + Total as formulas so the sheet stays live if cells are edited
    first = get_column_letter(label_cols + 1)
    last = get_column_letter(label_cols + len(exch_cols))
    tot_c, ven_c = len(matrix.columns) + 1, len(matrix.columns) + 2
    ws.cell(2, tot_c, "Total Volume")
    ws.cell(2, ven_c, "# Venues")
    for r in range(3, n_rows + 3):
        ws.cell(r, tot_c, f"=SUM({first}{r}:{last}{r})")
        ws.cell(r, ven_c, f"=COUNT({first}{r}:{last}{r})")

    for c in range(1, ven_c + 1):
        h = ws.cell(2, c)
        h.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        h.fill = HEADER_FILL
        h.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = (
            28 if c == 1 else 16 if c <= label_cols else 15)
    ws.row_dimensions[2].height = 34

    for r in range(3, n_rows + 3):
        for c in range(1, ven_c + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT, size=10)
            cell.border = Border(bottom=THIN)
            if c > label_cols:
                # blank = not listed there; that absence is itself a finding
                cell.number_format = '#,##0;(#,##0);"-"'
            elif c <= label_cols:
                cell.fill = GROUP_FILL
        ws.cell(r, tot_c).font = Font(name=FONT, size=10, bold=True)

    ws.freeze_panes = ws.cell(3, label_cols + 1)
    ws.auto_filter.ref = f"A2:{get_column_letter(ven_c)}{n_rows + 2}"

    d = wb["Contract Detail"]
    for c in range(1, d.max_column + 1):
        d.cell(1, c).font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        d.cell(1, c).fill = HEADER_FILL
        d.column_dimensions[get_column_letter(c)].width = 18
    d.freeze_panes = "A2"

    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("infile")
    ap.add_argument("--out", default="tradfi_matrix.xlsx")
    ap.add_argument("--merge-hl", action="store_true",
                    help="collapse Hyperliquid builder namespaces into one column")
    args = ap.parse_args()

    df = pd.read_csv(args.infile)
    matrix, exch_cols, detail = build(df, merge_hl=args.merge_hl)
    write(matrix, exch_cols, detail, args.out)

    print(f"{len(matrix)} underlyings x {len(exch_cols)} exchanges -> {args.out}")
    print("\nvenues per underlying (top 10 by coverage):")
    cov = matrix.set_index("Underlying")[exch_cols].notna().sum(axis=1)
    print(cov.sort_values(ascending=False).head(10).to_string())


if __name__ == "__main__":
    main()